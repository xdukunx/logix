#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import sqlite3
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Any

DEFAULT_DB = Path(os.environ.get("NOTIFY_DB", "/opt/software/notify/notify.db"))
DEFAULT_OUTDIR = Path(os.environ.get("LOGBOOK_REPORT_DIR", "/opt/software/notify/reports"))
WINDOWS_SESSION_JSON = Path(os.environ.get("LOGBOOK_SESSION_JSON", "/mnt/c/ProgramData/MindLabLogbook/session.json"))

BASE_COLUMNS = {
    "id": "INTEGER PRIMARY KEY AUTOINCREMENT",
    "timestamp": "TEXT NOT NULL",
    "event": "TEXT NOT NULL",
    "username": "TEXT",
    "nama": "TEXT",
    "nim": "TEXT",
    "tujuan": "TEXT",
    "keterangan": "TEXT",
    "session_type": "TEXT",
    "source": "TEXT",
    "session_id": "TEXT",
    "windows_user": "TEXT",
    "hostname": "TEXT",
    "client_ip": "TEXT",
    "anydesk_detected": "INTEGER DEFAULT 0",
    "raw_json": "TEXT",
}

CLOSE_EVENTS = {"END", "LOCK", "AUTO_FINISH", "AUTO_CLOSE", "DISCONNECT", "LOGOFF"}


def parse_date(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d")


def connect(db: Path) -> sqlite3.Connection:
    con = sqlite3.connect(str(db), timeout=30)
    con.row_factory = sqlite3.Row
    return con


def table_exists(con: sqlite3.Connection, table: str) -> bool:
    r = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return r is not None


def sqlite_columns(con: sqlite3.Connection, table: str) -> set[str]:
    try:
        return {str(r[1]) for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
    except sqlite3.Error:
        return set()


def _alter_decl_for_existing_table(decl: str) -> str:
    upper = decl.upper()
    if "PRIMARY KEY" in upper:
        return decl
    if "NOT NULL" in upper and "DEFAULT" not in upper:
        if upper.startswith("TEXT"):
            return decl.replace("NOT NULL", "").replace("not null", "").strip() + " DEFAULT ''"
        if upper.startswith("INTEGER"):
            return decl.replace("NOT NULL", "").replace("not null", "").strip() + " DEFAULT 0"
        return decl.replace("NOT NULL", "").replace("not null", "").strip()
    return decl


def ensure_physical_schema(con: sqlite3.Connection) -> None:
    """Create/repair physical_log schema before report-time active-session repair."""
    col_defs = ",\n        ".join(f"{k} {v}" for k, v in BASE_COLUMNS.items())
    con.execute(f"CREATE TABLE IF NOT EXISTS physical_log (\n        {col_defs}\n    )")
    cols = sqlite_columns(con, "physical_log")
    for name, decl in BASE_COLUMNS.items():
        if name not in cols and "PRIMARY KEY" not in decl.upper():
            con.execute(f"ALTER TABLE physical_log ADD COLUMN {name} {_alter_decl_for_existing_table(decl)}")
    con.execute("CREATE INDEX IF NOT EXISTS idx_physical_log_timestamp ON physical_log(timestamp)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_physical_log_session ON physical_log(session_id)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_physical_log_type ON physical_log(session_type)")
    con.execute("CREATE INDEX IF NOT EXISTS idx_physical_log_event ON physical_log(event)")
    con.commit()


def _norm(v: Any, default: str = "") -> str:
    if v is None:
        return default
    return str(v).strip()


def _truthy(v: Any) -> int:
    return 1 if str(v).strip().lower() in {"1", "true", "yes", "y", "anydesk"} else 0


def _active_payload_from_session_file(session_path: Path) -> dict[str, Any] | None:
    if not session_path.exists():
        return None
    try:
        data = json.loads(session_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return None
    sid = _norm(data.get("session_id"))
    if not sid:
        return None
    payload = {
        "timestamp": _report_ts_text(data.get("start_time") or data.get("timestamp"), fallback=datetime.now()),
        "event": "START",
        "username": _norm(data.get("username")),
        "nama": _norm(data.get("nama")),
        "nim": _norm(data.get("nim")),
        "tujuan": _norm(data.get("tujuan")),
        "keterangan": _norm(data.get("keterangan")),
        "session_type": _norm(data.get("session_type"), "Physical"),
        "source": "windows_session_json_repair",
        "session_id": sid,
        "windows_user": _norm(data.get("windows_user")),
        "hostname": _norm(data.get("hostname")),
        "client_ip": _norm(data.get("client_ip")),
        "anydesk_detected": _truthy(data.get("anydesk_detected")),
        "raw_json": json.dumps({"report_time_repair": True, "session_json": data}, ensure_ascii=False),
    }
    if not payload["username"] and payload["windows_user"]:
        payload["username"] = payload["windows_user"].split("\\")[-1]
    return payload


def repair_active_session_from_windows_state(con: sqlite3.Connection, session_path: Path | None = None) -> str:
    """Ensure the currently active Windows popup session is present in SQLite.

    This fixes the real-world failure mode where the WPF form creates
    C:\\ProgramData\\MindLabLogbook\\session.json and starts the timer, but the
    START insert into WSL/SQLite is missed because WSL, Python, or permission
    handling was restarted. Report generation is allowed to do this small
    idempotent repair before reading rows, so an active user is never invisible
    in /audit_report.
    """
    session_path = session_path or WINDOWS_SESSION_JSON
    payload = _active_payload_from_session_file(session_path)
    if not payload:
        return "no_active_session_json"
    ensure_physical_schema(con)
    cols = list(BASE_COLUMNS.keys())
    cols.remove("id")
    sid = payload["session_id"]
    existing = con.execute(
        """
        SELECT id FROM physical_log
         WHERE session_id=? AND UPPER(COALESCE(event,''))='START'
         ORDER BY id ASC LIMIT 1
        """,
        (sid,),
    ).fetchone()
    if existing:
        vals = dict(payload)
        vals["id"] = existing["id"]
        con.execute(
            """
            UPDATE physical_log
               SET timestamp = CASE WHEN COALESCE(timestamp,'')='' THEN :timestamp ELSE timestamp END,
                   username = CASE WHEN :username<>'' THEN :username ELSE username END,
                   nama = CASE WHEN :nama<>'' THEN :nama ELSE nama END,
                   nim = CASE WHEN :nim<>'' THEN :nim ELSE nim END,
                   tujuan = CASE WHEN :tujuan<>'' THEN :tujuan ELSE tujuan END,
                   keterangan = CASE WHEN :keterangan<>'' THEN :keterangan ELSE keterangan END,
                   session_type = CASE WHEN :session_type<>'' THEN :session_type ELSE session_type END,
                   source = CASE WHEN COALESCE(source,'')='' THEN :source ELSE source END,
                   windows_user = CASE WHEN :windows_user<>'' THEN :windows_user ELSE windows_user END,
                   hostname = CASE WHEN :hostname<>'' THEN :hostname ELSE hostname END,
                   client_ip = CASE WHEN :client_ip<>'' THEN :client_ip ELSE client_ip END,
                   anydesk_detected = :anydesk_detected,
                   raw_json = CASE WHEN COALESCE(raw_json,'')='' THEN :raw_json ELSE raw_json END
             WHERE id=:id
            """,
            vals,
        )
        con.commit()
        return f"updated_active_session:{sid}"

    placeholders = ",".join("?" for _ in cols)
    try:
        con.execute(
            f"INSERT INTO physical_log ({','.join(cols)}) VALUES ({placeholders})",
            [payload.get(c, "") for c in cols],
        )
        con.commit()
        return f"inserted_active_session:{sid}"
    except sqlite3.IntegrityError:
        # Legacy database may still have a bad UNIQUE(session_id). If insert fails,
        # update by session_id so report generation still succeeds.
        con.execute(
            """
            UPDATE physical_log
               SET timestamp=:timestamp, event='START', username=:username, nama=:nama, nim=:nim,
                   tujuan=:tujuan, keterangan=:keterangan, session_type=:session_type, source=:source,
                   windows_user=:windows_user, hostname=:hostname, client_ip=:client_ip,
                   anydesk_detected=:anydesk_detected, raw_json=:raw_json
             WHERE session_id=:session_id
            """,
            payload,
        )
        con.commit()
        return f"updated_active_session_after_integrity:{sid}"


def fetch_physical(con: sqlite3.Connection, start: str | None, end: str | None):
    if not table_exists(con, "physical_log"):
        return []
    q = "SELECT * FROM physical_log WHERE 1=1"
    args = []
    if start:
        q += " AND timestamp >= ?"
        args.append(start)
    if end:
        q += " AND timestamp < ?"
        args.append(end)
    q += " ORDER BY timestamp ASC, id ASC"
    return con.execute(q, args).fetchall()


def fetch_jobs(con: sqlite3.Connection, start: str | None, end: str | None):
    if not table_exists(con, "jobs"):
        return []
    cols = [r[1] for r in con.execute("PRAGMA table_info(jobs)").fetchall()]
    ts_col = next((c for c in ["started_at", "start_time", "created_at", "timestamp", "ts"] if c in cols), None)
    if not ts_col:
        return []
    q = f"SELECT {','.join(cols)} FROM jobs WHERE 1=1"
    args = []
    if start:
        q += f" AND {ts_col} >= ?"
        args.append(start)
    if end:
        q += f" AND {ts_col} < ?"
        args.append(end)
    q += f" ORDER BY {ts_col} ASC"
    try:
        return con.execute(q, args).fetchall()
    except sqlite3.Error:
        return []


def safe_get(row, key, default=""):
    try:
        return row[key] if row[key] is not None else default
    except Exception:
        return default


def first_nonempty(rows, key, default=""):
    for row in rows:
        if not row:
            continue
        val = safe_get(row, key)
        if str(val or "").strip():
            return val
    return default


def _strip_tzinfo_keep_wall_time(dt: datetime | None) -> datetime | None:
    """Return a timezone-naive datetime while preserving the displayed local clock time.

    Windows session.json may write ISO strings such as 2026-06-08T15:18:00+07:00,
    while older SQLite rows are usually stored without an offset.  Subtracting
    aware and naive datetimes raises TypeError.  For a local workstation logbook,
    the safest report behavior is to keep the wall-clock time that users see in
    Windows and strip only tzinfo for duration/report calculations.
    """
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt


def _report_ts_text(value: Any, fallback: datetime | None = None) -> str:
    """Normalize timestamp text before inserting report-time repaired rows."""
    dt = parse_ts(str(value)) if value is not None and str(value).strip() else fallback
    dt = _strip_tzinfo_keep_wall_time(dt)
    if dt:
        return dt.isoformat(timespec="seconds")
    return _norm(value)


def parse_ts(s: str | None):
    if not s:
        return None
    txt = str(s).strip().replace("Z", "+00:00")
    import re
    m = re.match(r"^(.*T\d{2}:\d{2}:\d{2})\.(\d{6})\d+(.*)$", txt)
    if m:
        txt = m.group(1) + "." + m.group(2) + m.group(3)
    try:
        return _strip_tzinfo_keep_wall_time(datetime.fromisoformat(txt))
    except Exception:
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return _strip_tzinfo_keep_wall_time(datetime.strptime(txt[:19], fmt))
            except Exception:
                pass
        return None


def fmt_ts(s: str | None) -> str:
    dt = parse_ts(s)
    if dt:
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(s or "")


def fmt_duration(start_s: str | None, end_s: str | None, active: bool = False) -> str:
    start = parse_ts(start_s)
    end = parse_ts(end_s) if end_s else None
    if active and start and not end:
        end = datetime.now(start.tzinfo) if start.tzinfo else datetime.now()
    if not start or not end:
        return "-"
    seconds = max(0, int((end - start).total_seconds()))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h:
        return f"{h}j {m}m {s}d"
    if m:
        return f"{m}m {s}d"
    return f"{s}d"


def event_priority(row):
    event = (safe_get(row, "event") or "").upper()
    if event == "START":
        return 0
    if event in CLOSE_EVENTS:
        return 1
    return 2


def build_sessions(rows):
    sessions = {}
    loose = []
    for row in rows:
        event = (safe_get(row, "event") or "").upper()
        sid = safe_get(row, "session_id") or ""
        if sid and (event == "START" or event in CLOSE_EVENTS or event in {"UNLOCK", "SSH_LOGIN", "SSH_LOGOUT"}):
            rec = sessions.setdefault(sid, {"starts": [], "ends": [], "events": []})
            rec["events"].append(row)
            if event == "START":
                rec["starts"].append(row)
            elif event in CLOSE_EVENTS:
                rec["ends"].append(row)
            else:
                # SSH_LOGIN without START remains loose-like, but grouped by sid if possible.
                if event == "SSH_LOGIN" and not rec["starts"]:
                    rec["starts"].append(row)
                elif event == "SSH_LOGOUT":
                    rec["ends"].append(row)
        else:
            loose.append(row)

    out = []
    for sid, rec in sessions.items():
        starts = sorted(rec["starts"], key=lambda r: (safe_get(r, "timestamp") or "", safe_get(r, "id") or 0))
        ends = sorted(rec["ends"], key=lambda r: (safe_get(r, "timestamp") or "", safe_get(r, "id") or 0))
        ordered_events = sorted(rec["events"], key=lambda r: (event_priority(r), safe_get(r, "timestamp") or "", safe_get(r, "id") or 0))
        start = starts[0] if starts else (ordered_events[0] if ordered_events else None)
        # Prefer START row with identity filled for display, because older broken builds inserted a blank START first.
        rich_start = next((r for r in starts if safe_get(r, "nama") or safe_get(r, "nim") or safe_get(r, "tujuan")), start)
        end = ends[-1] if ends else None
        if not start:
            continue
        display_rows = [rich_start, start] + ordered_events + ([end] if end else [])
        stype = first_nonempty(display_rows, "session_type")
        event_end = (safe_get(end, "event") or "").upper() if end else ""
        status = "Selesai / Finish" if end else "Aktif"
        if event_end in {"AUTO_FINISH", "AUTO_CLOSE"}:
            status = "Auto Finish"
        kategori = "Remote" if str(stype).upper() == "SSH" else "Workstation"
        out.append({
            "start_ts": safe_get(start, "timestamp"),
            "end_ts": safe_get(end, "timestamp") if end else "",
            "nama": first_nonempty(display_rows, "nama") or first_nonempty(display_rows, "username") or first_nonempty(display_rows, "windows_user"),
            "nim": first_nonempty(display_rows, "nim"),
            "tujuan": first_nonempty(display_rows, "tujuan"),
            "tipe": stype,
            "status": status,
            "durasi": fmt_duration(safe_get(start, "timestamp"), safe_get(end, "timestamp") if end else None, active=(not end)),
            "keterangan": first_nonempty(display_rows, "keterangan"),
            "kategori": kategori,
            "session_id": sid,
            "_active": not bool(end),
        })

    for row in loose:
        event = (safe_get(row, "event") or "").upper()
        stype = safe_get(row, "session_type") or ("SSH" if event.startswith("SSH") else "")
        out.append({
            "start_ts": safe_get(row, "timestamp"),
            "end_ts": "",
            "nama": safe_get(row, "nama") or safe_get(row, "username") or safe_get(row, "windows_user"),
            "nim": safe_get(row, "nim"),
            "tujuan": safe_get(row, "tujuan") or event,
            "tipe": stype,
            "status": "Tercatat",
            "durasi": "-",
            "keterangan": safe_get(row, "keterangan") or safe_get(row, "client_ip") or safe_get(row, "hostname"),
            "kategori": "Remote" if str(stype).upper() == "SSH" else "Workstation",
            "session_id": safe_get(row, "session_id"),
            "_active": False,
        })

    out.sort(key=lambda r: r.get("start_ts") or "")

    # Presentation guard for legacy data: if multiple workstation sessions are active, only the newest remains Aktif.
    active_idx = [i for i, r in enumerate(out) if r.get("_active") and str(r.get("kategori")).upper() == "WORKSTATION"]
    for pos, idx in enumerate(active_idx[:-1]):
        next_idx = active_idx[pos + 1]
        out[idx]["end_ts"] = out[next_idx].get("start_ts") or ""
        out[idx]["status"] = "Auto Finish"
        out[idx]["durasi"] = fmt_duration(out[idx].get("start_ts"), out[idx].get("end_ts"), active=False)
        out[idx]["_active"] = False

    return out


def write_xlsx(rows, jobs, output: Path, period_label: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise SystemExit("openpyxl belum tersedia untuk Python yang menjalankan logbook_report.py. Jalankan sudo bash update.sh dari paket ini.") from exc

    sessions = build_sessions(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report Logbook"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    dark = PatternFill("solid", fgColor="0B1220")
    teal = PatternFill("solid", fgColor="14B8A6")
    header = PatternFill("solid", fgColor="E5E7EB")
    white = PatternFill("solid", fgColor="FFFFFF")
    active_fill = PatternFill("solid", fgColor="ECFDF5")
    finish_fill = PatternFill("solid", fgColor="F8FAFC")
    auto_fill = PatternFill("solid", fgColor="FEF3C7")
    remote_fill = PatternFill("solid", fgColor="DBEAFE")
    physical_fill = PatternFill("solid", fgColor="F1F5F9")
    anydesk_fill = PatternFill("solid", fgColor="EDE9FE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    ws["A1"] = "MINDLAB  |  REPORT LOGBOOK"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].fill = dark
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Periode: {period_label} | Dibuat: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(color="CCFBF1", italic=True)
    ws["A2"].fill = dark
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:K3")
    ws["A3"] = "Physical / AnyDesk dicatat dari Windows popup. SSH/job remote dicatat dari WSL/notify-run."
    ws["A3"].font = Font(color="111827", italic=True, size=11)
    ws["A3"].fill = PatternFill("solid", fgColor="F0FDFA")
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = ["No", "Waktu Mulai", "Waktu Selesai", "Nama / User", "NIM / ID", "Tujuan", "Tipe Akses", "Status", "Durasi", "Keterangan", "Kategori"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = Font(color="111827", bold=True)
        c.fill = header
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, row in enumerate(sessions, 1):
        r = 4 + i
        values = [
            i,
            fmt_ts(row["start_ts"]),
            fmt_ts(row["end_ts"]),
            row["nama"],
            row["nim"],
            row["tujuan"],
            row["tipe"],
            row["status"],
            row["durasi"],
            row["keterangan"],
            row["kategori"],
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(r, col, val)
            c.fill = white
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 7:
                access = str(val).upper()
                c.fill = {"ANYDESK": anydesk_fill, "PHYSICAL": physical_fill, "SSH": remote_fill}.get(access, white)
                c.font = Font(bold=True, color="111827")
            if col == 8:
                status = str(val).upper()
                if "AKTIF" in status:
                    c.fill = active_fill
                elif "AUTO" in status:
                    c.fill = auto_fill
                elif "FINISH" in status or "SELESAI" in status:
                    c.fill = finish_fill
                c.font = Font(bold=True, color="111827")

    last = max(5, 4 + len(sessions))
    ws.auto_filter.ref = f"A4:K{last}"
    widths = [6, 21, 21, 26, 16, 34, 14, 16, 14, 46, 16]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm["A1"] = "Ringkasan Logbook"
    sm["A1"].font = Font(bold=True, size=15, color="111827")
    sm.append(["Metrik", "Nilai"])
    active = sum(1 for r in sessions if "Aktif" in r["status"])
    finished = sum(1 for r in sessions if "Finish" in r["status"] or "Selesai" in r["status"])
    auto_finish = sum(1 for r in sessions if "Auto" in r["status"])
    ssh = sum(1 for r in sessions if str(r["tipe"]).upper() == "SSH")
    anydesk = sum(1 for r in sessions if str(r["tipe"]).upper() == "ANYDESK")
    physical = sum(1 for r in sessions if str(r["tipe"]).upper() == "PHYSICAL")
    for item in [
        ("Total sesi/event logbook", len(sessions)),
        ("Sesi aktif", active),
        ("Selesai / Finish", finished),
        ("Auto Finish (legacy/stale)", auto_finish),
        ("Physical", physical),
        ("AnyDesk", anydesk),
        ("SSH", ssh),
        ("Job terdeteksi", len(jobs)),
    ]:
        sm.append(list(item))
    for row in sm.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    for cell in sm[2]:
        cell.fill = header
        cell.font = Font(bold=True)
    sm.column_dimensions["A"].width = 32
    sm.column_dimensions["B"].width = 20

    js = wb.create_sheet("Jobs")
    js.sheet_view.showGridLines = False
    if jobs:
        cols = list(jobs[0].keys())
        js.append(cols)
        for cell in js[1]:
            cell.fill = header
            cell.font = Font(color="111827", bold=True)
            cell.border = border
        for row in jobs:
            js.append([safe_get(row, c) for c in cols])
        for row in js.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx, col in enumerate(cols, 1):
            js.column_dimensions[get_column_letter(idx)].width = min(max(len(col) + 4, 14), 36)
        js.auto_filter.ref = js.dimensions
        js.freeze_panes = "A2"
    else:
        js["A1"] = "Tidak ada tabel jobs atau tidak ada job pada periode ini."

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)




def _coerce_report_date(value: Any) -> date | None:
    """Normalize date/datetime/string values accepted by build()."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        normalized = txt.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(normalized).date()
        except ValueError:
            return datetime.strptime(txt[:10], "%Y-%m-%d").date()
    raise TypeError("tanggal harus None, str YYYY-MM-DD, datetime.date, atau datetime.datetime")


def _coerce_report_datetime(value: Any) -> datetime | None:
    """Normalize optional exact datetime values accepted by build(start_dt/end_dt)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return _strip_tzinfo_keep_wall_time(value)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return None
        normalized = txt.replace("Z", "+00:00")
        try:
            return _strip_tzinfo_keep_wall_time(datetime.fromisoformat(normalized))
        except ValueError:
            try:
                return datetime.strptime(txt[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return datetime.strptime(txt[:10], "%Y-%m-%d")
    raise TypeError("start_dt/end_dt harus None, str, datetime.date, atau datetime.datetime")


def _normalize_build_aliases(start_date: Any, end_date: Any, start_dt: Any, end_dt: Any, kwargs: dict[str, Any]):
    """
    Accept old/new caller names.

    Telegram bot builds found in the wild call build(start_dt=..., end_dt=...),
    while the CLI wrapper uses start_date/end_date.  This normalizer keeps both
    contracts valid so future bot versions do not break report generation again.
    """
    if "start" in kwargs and start_date is None:
        start_date = kwargs.pop("start")
    if "end" in kwargs and end_date is None:
        end_date = kwargs.pop("end")
    if "from_date" in kwargs and start_date is None:
        start_date = kwargs.pop("from_date")
    if "to_date" in kwargs and end_date is None:
        end_date = kwargs.pop("to_date")
    if "date_from" in kwargs and start_date is None:
        start_date = kwargs.pop("date_from")
    if "date_to" in kwargs and end_date is None:
        end_date = kwargs.pop("date_to")
    if "start_time" in kwargs and start_dt is None:
        start_dt = kwargs.pop("start_time")
    if "end_time" in kwargs and end_dt is None:
        end_dt = kwargs.pop("end_time")
    return start_date, end_date, start_dt, end_dt


def _resolve_report_period(start_date: Any = None, end_date: Any = None, full: bool = False,
                           start_dt: Any = None, end_dt: Any = None):
    """Return (start_filter, end_filter, label, filename_suffix) for report generation."""
    if full:
        return None, None, "Semua data", "full"

    # Exact datetime mode, used by some Telegram bot handlers.
    start_exact = _coerce_report_datetime(start_dt)
    end_exact = _coerce_report_datetime(end_dt)
    if start_exact is not None or end_exact is not None:
        if start_exact is None and end_exact is not None:
            start_exact = datetime.combine(end_exact.date(), time.min)
        elif start_exact is not None and end_exact is None:
            end_exact = start_exact + timedelta(days=1)
        if end_exact < start_exact:
            raise ValueError("end_dt tidak boleh lebih awal dari start_dt")

        # Bot lama mengirim end_dt pada pukul 00:00 tanggal akhir, misalnya
        # 2026-06-08 00:00. Kalau ini dipakai sebagai batas eksklusif, semua log
        # pada 2026-06-08 pukul 00:01-23:59 akan hilang dan Excel terlihat kosong.
        # Karena itu, midnight date boundary yang tidak berada di masa depan
        # diperlakukan sebagai tanggal akhir inklusif. Kalau caller sudah mengirim
        # besok 00:00 sebagai exclusive boundary, nilainya tidak diubah.
        end_filter_exact = end_exact
        end_is_midnight_boundary = (end_exact.time() == time.min)
        if end_is_midnight_boundary and end_exact.date() <= date.today():
            end_filter_exact = end_exact + timedelta(days=1)

        start_s = start_exact.isoformat(timespec="seconds")
        end_s = end_filter_exact.isoformat(timespec="seconds")

        if start_exact.time() == time.min and end_exact.time() == time.min:
            label = start_exact.date().isoformat() if start_exact.date() == end_exact.date() else f"{start_exact.date().isoformat()} s.d. {end_exact.date().isoformat()}"
            suffix = start_exact.strftime("%Y%m%d") if start_exact.date() == end_exact.date() else f"{start_exact.strftime('%Y%m%d')}-{end_exact.strftime('%Y%m%d')}"
        elif start_exact.date() == end_exact.date():
            label = start_exact.date().isoformat()
            suffix = start_exact.strftime("%Y%m%d")
        else:
            label = f"{start_exact.strftime('%Y-%m-%d %H:%M')} s.d. {end_exact.strftime('%Y-%m-%d %H:%M')}"
            suffix = f"{start_exact.strftime('%Y%m%d-%H%M')}-{end_exact.strftime('%Y%m%d-%H%M')}"
        return start_s, end_s, label, suffix

    start_d = _coerce_report_date(start_date)
    end_d = _coerce_report_date(end_date)

    # Default wrapper untuk bot Telegram: report harian jika rentang tidak diberikan.
    if start_d is None and end_d is None:
        start_d = end_d = date.today()
    elif start_d is None and end_d is not None:
        # Jika hanya end_date dikirim dari modul lain, jadikan single-day report pada tanggal tersebut.
        start_d = end_d
    elif start_d is not None and end_d is None:
        end_d = start_d

    if end_d < start_d:
        raise ValueError("end_date tidak boleh lebih awal dari start_date")

    start_filter_dt = datetime.combine(start_d, time.min)
    end_excl = datetime.combine(end_d, time.min) + timedelta(days=1)

    start_filter = start_filter_dt.isoformat(timespec="seconds")
    end_filter = end_excl.isoformat(timespec="seconds")
    label = start_d.isoformat() if start_d == end_d else f"{start_d.isoformat()} s.d. {end_d.isoformat()}"
    suffix = start_d.strftime("%Y%m%d") if start_d == end_d else f"{start_d.strftime('%Y%m%d')}-{end_d.strftime('%Y%m%d')}"
    return start_filter, end_filter, label, suffix


def build(start_date: Any = None,
          end_date: Any = None,
          out_dir: str | os.PathLike[str] | None = None,
          full: bool = False,
          start_dt: Any = None,
          end_dt: Any = None,
          db: str | os.PathLike[str] | None = None,
          db_path: str | os.PathLike[str] | None = None,
          output: str | os.PathLike[str] | None = None,
          out: str | os.PathLike[str] | None = None,
          repair_active: bool | None = None,
          **kwargs: Any) -> str:
    """
    Wrapper report generator untuk dipanggil modul lain, termasuk bot Telegram.

    Kompatibel dengan beberapa versi caller:
    - build(start_date=..., end_date=..., out_dir=..., full=False)
    - build(start_dt=..., end_dt=..., out_dir=..., full=False)
    - build(start=..., end=..., output=...)

    Returns
    -------
    str
        Path file .xlsx yang berhasil dibuat.
    """
    start_date, end_date, start_dt, end_dt = _normalize_build_aliases(
        start_date, end_date, start_dt, end_dt, kwargs
    )

    # Abaikan keyword tambahan yang tidak penting supaya bot lama tidak crash.
    # Contoh yang pernah muncul: chat_id, user_id, period, mode.
    start_s, end_s, period_label, suffix = _resolve_report_period(
        start_date=start_date,
        end_date=end_date,
        full=full,
        start_dt=start_dt,
        end_dt=end_dt,
    )

    db_file = Path(db_path or db or DEFAULT_DB).expanduser()

    explicit_output = output or out
    if explicit_output:
        output_path = Path(explicit_output).expanduser()
        output_path.parent.mkdir(parents=True, exist_ok=True)
    else:
        output_dir = Path(out_dir) if out_dir is not None else DEFAULT_OUTDIR
        output_dir = output_dir.expanduser()
        output_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output_path = output_dir / f"report-logbook-{suffix}-{stamp}.xlsx"

    con = connect(db_file)
    try:
        # Idempotent repair: if the popup/timer has an active session but SQLite
        # missed its START row, insert/update it before the report query.
        # Do this automatically only for the real/default DB so test/custom DBs are
        # not polluted by the workstation's live Windows session.json.
        if repair_active is None:
            try:
                repair_active = db_file.resolve() == DEFAULT_DB.expanduser().resolve()
            except Exception:
                repair_active = str(db_file) == str(DEFAULT_DB.expanduser())
        if repair_active:
            repair_active_session_from_windows_state(con)
        rows = fetch_physical(con, start_s, end_s)
        jobs = fetch_jobs(con, start_s, end_s)
        write_xlsx(rows, jobs, output_path, period_label)
    finally:
        con.close()

    return str(output_path)

def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Generate Excel Report Logbook")
    p.add_argument("--db", default=str(DEFAULT_DB))
    p.add_argument("--out", default="")
    p.add_argument("--start", default="", help="YYYY-MM-DD")
    p.add_argument("--end", default="", help="YYYY-MM-DD inclusive")
    p.add_argument("--today", action="store_true")
    p.add_argument("--full", action="store_true")
    ns = p.parse_args(argv)

    start_dt = end_excl = None
    label = "Semua data"
    if ns.today:
        d = datetime.combine(date.today(), time.min)
        start_dt = d
        end_excl = d + timedelta(days=1)
        label = date.today().isoformat()
    elif ns.start or ns.end:
        if not ns.start:
            raise SystemExit("--start wajib kalau --end dipakai")
        start_dt = parse_date(ns.start)
        if ns.end:
            end_excl = parse_date(ns.end) + timedelta(days=1)
            label = f"{ns.start} s.d. {ns.end}"
        else:
            end_excl = start_dt + timedelta(days=1)
            label = ns.start

    start_s = start_dt.isoformat(timespec="seconds") if start_dt else None
    end_s = end_excl.isoformat(timespec="seconds") if end_excl else None

    out = Path(ns.out) if ns.out else DEFAULT_OUTDIR / f"report-logbook-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    con = connect(Path(ns.db))
    try:
        repair_active_session_from_windows_state(con)
        rows = fetch_physical(con, start_s, end_s)
        jobs = fetch_jobs(con, start_s, end_s)
        write_xlsx(rows, jobs, out, label)
    finally:
        con.close()
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(__import__('sys').argv[1:]))
